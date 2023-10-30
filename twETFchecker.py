#! python3
#==============================================================================
#    Copyright (c) 2023 JimmyLi. All rights reserved.
#    This program contains proprietary and confidential information.
#    All rights reserved except as may be permitted by prior written consent.
#
#
#    ModuleName:
#            ETFWatcher.py
#
#    Abstract:
#            get target ETF stock structure and the prize
#
#    Author:
#            02-Aug-2023 Jimmy Li
#
#    Revision History:
#           Rev  1.0.2 Jimmy Li
#               adjust
#                   1,getTop20 : the method of store top20
#           Rev  1.0.1 Jimmy Li
#                First create.
#==============================================================================
import openpyxl
import sys
import logging
import os
from datetime import datetime
from logging import handlers
from selenium import webdriver
from selenium.webdriver.common.by import By

class ETFwatcher():
    def __init__(self):
        self.strFailReason = ""
        self.logger = self.init_logger()
        self.strETFExcelName = "testETF總表.xlsx"
        self.strETFsheet = "ETF"
        self.nStockNumberColumn  = 1
        self.dictStockPosition = {}
        self.url = ""
        self.listStockTemp = []
        self.dictStock = {}
        self.nStockInfoColumn = 2
        self.sheet = None
        
    def init_logger(self):
        if not os.path.exists('logs'):
            os.makedirs("logs")
        log_filename = datetime.now().strftime("logs/ETFwatcher-%Y-%m-%d.log")
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.DEBUG)
        formatter = logging.Formatter("%(asctime)s - [line:%(lineno)d] - %(name)s - %(levelname)s: %(message)s")
        file_handler = handlers.TimedRotatingFileHandler(filename=log_filename, when='midnight', backupCount=30, encoding='utf-8')
        file_handler.setFormatter(formatter)
        file_handler.setLevel(logging.DEBUG)
        logger.addHandler(file_handler)
        return logger
    
    def readETFExcel(self):
        self.logger.info("====== start read ETFExcel ======")
        try:

            self.logger.info("excel exist %s" ,str(os.path.isfile(self.strETFExcelName)))
            self.workbook = openpyxl.load_workbook(self.strETFExcelName)
            self.sheet = self.workbook[self.strETFsheet]
            self.logger.info("Excel load success")
        except:
            strFailReason = f"loadWorkbook fail (%s)" % (str(sys.exc_info()[1]))

            return [False, strFailReason]

        self.logger.info(f"get sheet: {self.sheet}")
        return [True, ""]

    def getStockNumber(self):
        self.logger.info("===== start getStockNumber =====")

        # read without ColumnName
        for nRow in range(2, self.sheet.max_row + 1):
            cell = self.sheet.cell(nRow, self.nStockNumberColumn)
            if cell.value != "" and cell.value != None:
                self.dictStockPosition[cell.value] = nRow
                print(cell.value, nRow)
        self.logger.info(f"get StockerNumber success StockNumber and coordination : {self.dictStockPosition}")

    def getStockTodayInfo(self, nStockNumber):
        self.logger.info("===== start getStockTodayInfo ======")
        try:
            self.browser.get(self.getRequestsUrl(nStockNumber))
            self.browser.implicitly_wait(10)
            self.strStockPrize = self.browser.find_element(By.XPATH, '//*[@id="__layout"]/div/div[3]/div/div[2]/main/div/div[1]/div/div/article/div[1]/div[2]/div[1]/span[1]').text
            self.bs4StockStructure = self.browser.find_element(By.XPATH, '//*[@id="__layout"]/div/div[3]/div/div[2]/main/div/div[4]/section/div[2]/div/table/tbody')
            print(f"strStockPrize {self.strStockPrize}")
            self.logger.info(f"{nStockNumber} Prize {self.strStockPrize}")
            return [True, ""]
        except:
            strFailReason = "Exception Fail.(%s)" % (str(sys.exc_info()[1]))
            return [False, strFailReason]

    def getRequestsUrl(self, nStockNumber: int) -> str:
        url = f"https://www.cmoney.tw/etf/tw/{nStockNumber}/fundholding"
        self.logger.info(f"get url : {url}")
        return url
    
    def startBrower(self):
        self.logger.info("===== start startBrowser =====")
        self.browser = webdriver.Chrome()
        self.logger.info("start browser succsee")

# [1.0.1 10-Aug-2023 Jimmy Li] =>
    def writeNewStructure(self, nStockerNumber):
        self.logger.info("===== start writeNewStructure =====")
        try:
            self.sheet.cell(self.dictStockPosition[nStockerNumber], self.nStockInfoColumn).value = self.strStockPrize
            for dictStock in self.listTop20:
                col = 3
                for strKeyName in dictStock.keys():
                    self.sheet.cell(self.dictStockPosition[nStockerNumber], col).value = dictStock[strKeyName]
                    col = col + 1
                self.dictStockPosition[nStockerNumber] = self.dictStockPosition[nStockerNumber] + 1
        except:
            strFailReason = "Exception Fail.(%s)" % (str(sys.exc_info()[1]))
            return [False, strFailReason]
        self.logger.info("success writeNewStructure ")
        return [True, ""]

    def getTop20(self):
        self.logger.info("===== start getTop20 =====")
        print("start getTop20")
        listTop20Temp = []
        try:
            listStockStructure = self.bs4StockStructure.text.split()
            self.logger.info(listStockStructure)
        except:
            strFailReason = "listStockStructure split fail"
            return [False, strFailReason]

        # [1.0.2 19-Aug-2023 Jimmy Li] =>
        for nListIndex in range(0, len(listStockStructure)):
            if "%" in listStockStructure[nListIndex]:
                listTop20Temp.append({'StockNumber':listStockStructure[nListIndex-2],
                                    'StockName':listStockStructure[nListIndex-1],
                                    'Weights':listStockStructure[nListIndex],
                                    'Amount':listStockStructure[nListIndex+1],
                                    'Unit':listStockStructure[nListIndex+2]
                                    })
        # for Index in range(2, len(listStockStructure), 5):
        #     listTop20Temp.append({'StockNumber':listStockStructure[Index-2],
        #                         'StockName':listStockStructure[Index-1],
        #                         'Weights':listStockStructure[Index],
        #                         'Amount':listStockStructure[Index+1],
        #                         'Unit':listStockStructure[Index+2]
        #                         })
        # <= [1.0.2 19-Aug-2023 Jimmy Li]

        # make list sort by dict object of Weight key and get top20
        try:
            self.listTop20 = sorted(listTop20Temp, key = lambda i: i['Weights'],reverse=True)[:20]
        except:
            strFailReason = "listTop20 sort and get top20 fail"
            return [False, strFailReason]
        return [True, ""]
# <= [1.0.1 10-Aug-2023 Jimmy Li]

if __name__ == "__main__":

    cETF = ETFwatcher()

    cETF.logger.info(cETF.strETFExcelName)
    
    bRes, strFailReason = cETF.readETFExcel()

    if bRes is False:
        cETF.logger.error(f"readETFExcel error {strFailReason}")
        print(strFailReason)

    cETF.getStockNumber()

    cETF.startBrower()
    
    for nStockNumber in cETF.dictStockPosition.keys():
        cETF.logger.info(f"start search ETF stock: {nStockNumber}")
        cETF.logger.info(f"start search ETF stock: {nStockNumber}")
        cETF.bs4StockStructure = None
        bRes, strFailReason = cETF.getStockTodayInfo(nStockNumber)
        if bRes is False:
            cETF.logger.error(f"getStockTodayInfo error {strFailReason}")
            continue

        bRes, strFailReason = cETF.getTop20()
        if bRes is False:
            cETF.logger.error(f"getTop20 error {strFailReason}")
            continue
        
        bRes, strFailReason = cETF.writeNewStructure(nStockNumber)
        if bRes is False:
            cETF.logger.error(f"writeNewStructure error {strFailReason}")
            continue

        # 1.1.1 TODO make compaire
        # cETF.compaireWithOldInfo()
    # 1.1.1 TODO writeStructureChange()
    # cETF.writeStructureChange()

    cETF.workbook.save("test" + cETF.strETFExcelName)
    cETF.workbook.close()
    cETF.logger.info("===== process success =====")
    cETF.logger.info("===== process success =====")